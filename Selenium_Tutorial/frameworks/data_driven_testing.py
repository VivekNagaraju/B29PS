import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

filename=r"C:\Users\admin\Desktop\vivek.xlsx"
my_workbook=openpyxl.load_workbook(filename)
my_active_sheet=my_workbook.active
rows=my_active_sheet.max_row
columns=my_active_sheet.max_column
print(rows)
print(columns)

for r in range(2, rows+1):
    username_txt=my_active_sheet.cell(row=r, column=1).value
    print(username_txt)
    password_txt=my_active_sheet.cell(row=r, column=2).value
    print(password_txt)
    
    '''1. Launch Chrome Browser'''
    my_options=webdriver.ChromeOptions()
    my_options.add_experimental_option("detach", True)
    my_options.add_argument("start-maximized")
    driver=webdriver.Chrome(options=my_options)
    driver.implicitly_wait(5)
    
    '''2. Navigate to practice site'''
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    
    '''3. Enter UserName'''
    user_name=driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input')
    user_name.send_keys(username_txt)
    
    '''4. Enter Password'''
    password=driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[2]/div/div[2]/input')
    password.send_keys(password_txt)
    
    '''5. Click on login button'''
    login=driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button')
    login.click()
    
    '''6. Validate the login functionality'''
    actual_url=driver.current_url
    expected_url=my_active_sheet.cell(row=r, column=3).value
    # print(actual_url)
    
    if expected_url in actual_url:
        my_active_sheet.cell(row=r, column=4).value="Passed"
        my_workbook.save(filename)
        
        
    else:
        my_active_sheet.cell(row=r, column=4).value="Failed"
        my_workbook.save(filename)
    
        
