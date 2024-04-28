import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

filename=r"C:\Users\admin\Desktop\vivek.xlsx"
my_workbook=openpyxl.load_workbook(filename)
my_active_sheet=my_workbook.active
rows=my_active_sheet.max_row
columns=my_active_sheet.max_column
print(rows)
print(columns)
for r in range(2, rows+1):
    username_txt=my_active_sheet.cell(row=r, column=1).value
    print(username_txt)
    password_txt=my_active_sheet.cell(row=r, column=2).value
    print(password_txt)